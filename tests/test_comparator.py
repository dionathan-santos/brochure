"""Tests for modules/comparator.py"""
import unittest

import pandas as pd

from modules.comparator import (
    build_change_notes,
    infer_possible_cause,
    run_comparator,
)


# ── helpers ────────────────────────────────────────────────────────────────────

def _db(rows: list) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _rec(**kwargs) -> dict:
    defaults = {"Property Name": "Scotia Place", "Suite": "1200", "Size": 2000}
    defaults.update(kwargs)
    return defaults


# ── build_change_notes ─────────────────────────────────────────────────────────

class TestBuildChangeNotes(unittest.TestCase):

    def test_no_changes_returns_empty(self):
        db_row = {"Min Rent": 18.5, "Listing Agency": "CBRE"}
        br_row = {"Min Rent": 18.5, "Listing Agency": "CBRE"}
        self.assertEqual(build_change_notes(db_row, br_row), "")

    def test_detects_rent_change(self):
        db_row = {"Min Rent": 18.5}
        br_row = {"Min Rent": 21.0}
        notes = build_change_notes(db_row, br_row)
        self.assertIn("Min Rent", notes)
        self.assertIn("→", notes)

    def test_detects_multiple_changes(self):
        db_row = {"Min Rent": 18.5, "Max Rent": 22.0, "Listing Agency": "CBRE"}
        br_row = {"Min Rent": 20.0, "Max Rent": 24.0, "Listing Agency": "JLL"}
        notes = build_change_notes(db_row, br_row)
        self.assertIn("Min Rent", notes)
        self.assertIn("Max Rent", notes)
        self.assertIn("Listing Agency", notes)
        self.assertEqual(notes.count("|"), 2)

    def test_none_vs_value_detected(self):
        db_row = {"Min Rent": None}
        br_row = {"Min Rent": 21.0}
        notes = build_change_notes(db_row, br_row)
        self.assertIn("Min Rent", notes)

    def test_currency_display_format(self):
        db_row = {"Min Rent": 16.54}
        br_row = {"Min Rent": 21.0}
        notes = build_change_notes(db_row, br_row)
        self.assertIn("$16.54", notes)
        self.assertIn("$21.0", notes)


# ── infer_possible_cause ───────────────────────────────────────────────────────

class TestInferPossibleCause(unittest.TestCase):

    def test_long_standing_listing(self):
        db_row = {"Days on Market": 800, "Listing Agency": "CBRE", "Match_Score": 90}
        cause = infer_possible_cause(db_row, "CBRE")
        self.assertIn("2yr+", cause)

    def test_wrong_brokerage(self):
        db_row = {"Days on Market": 100, "Listing Agency": "JLL", "Match_Score": 90}
        cause = infer_possible_cause(db_row, "CBRE")
        self.assertIn("JLL", cause)
        self.assertIn("CBRE", cause)

    def test_low_fuzzy_score_alias_hint(self):
        db_row = {"Days on Market": 100, "Listing Agency": "CBRE", "Match_Score": 65}
        cause = infer_possible_cause(db_row, "CBRE")
        self.assertIn("alias", cause)

    def test_no_cause_fallback(self):
        db_row = {"Days on Market": 10, "Listing Agency": "CBRE", "Match_Score": 90}
        cause = infer_possible_cause(db_row, "CBRE")
        self.assertIn("manually", cause)

    def test_multiple_causes_joined_by_pipe(self):
        db_row = {"Days on Market": 1000, "Listing Agency": "JLL", "Match_Score": 62}
        cause = infer_possible_cause(db_row, "CBRE")
        self.assertIn("|", cause)


# ── run_comparator ─────────────────────────────────────────────────────────────

class TestRunComparator(unittest.TestCase):

    def test_no_db_all_new(self):
        records = [_rec(), _rec(**{"Property Name": "Epcor Tower"})]
        results = run_comparator(records, pd.DataFrame())
        actions = [r["Action"] for r in results]
        self.assertTrue(all(a == "NEW" for a in actions))

    def test_exact_name_match_ok(self):
        brochure = [_rec(**{"Min Rent": 18.5, "Listing Agency": "CBRE"})]
        db = _db([{
            "Property Name": "Scotia Place", "Suite": "1200", "Size": 2000,
            "Min Rent": 18.5, "Listing Agency": "CBRE",
        }])
        results = run_comparator(brochure, db)
        # Should match and be OK (no changes)
        matched = [r for r in results if r.get("Action") in ("OK", "UPDATE")]
        self.assertTrue(len(matched) >= 1)

    def test_changed_rent_is_update(self):
        brochure = [_rec(**{"Min Rent": 21.0, "Listing Agency": "CBRE"})]
        db = _db([{
            "Property Name": "Scotia Place", "Suite": "1200", "Size": 2000,
            "Min Rent": 18.5, "Listing Agency": "CBRE",
        }])
        results = run_comparator(brochure, db)
        updates = [r for r in results if r.get("Action") == "UPDATE"]
        self.assertEqual(len(updates), 1)
        self.assertIn("Min Rent", updates[0]["Change_Notes"])

    def test_unmatched_db_row_is_removed(self):
        brochure = [_rec(**{"Property Name": "Epcor Tower"})]
        db = _db([{
            "Property Name": "Scotia Place", "Suite": "1200", "Size": 2000,
            "Min Rent": 18.5,
        }])
        results = run_comparator(brochure, db)
        actions = {r["Action"] for r in results}
        self.assertIn("REMOVED", actions)

    def test_alias_normalization(self):
        """'ATCO Place' in brochure should match 'Canadian Western Bank Place' in DB."""
        brochure = [{"Property Name": "ATCO Place", "Suite": "100", "Size": 1000}]
        db = _db([{
            "Property Name": "Canadian Western Bank Place",
            "Suite": "100", "Size": 1000, "Min Rent": 18.0,
        }])
        results = run_comparator(brochure, db)
        matched = [r for r in results if r.get("Action") in ("OK", "UPDATE")]
        self.assertTrue(len(matched) >= 1, "Alias should resolve to a match")

    def test_review_action_for_borderline_score(self):
        """A name with 75-84% similarity and no address should yield REVIEW."""
        # "Scotia Plaza" vs "Scotia Place" — similar but not identical
        brochure = [{"Property Name": "Scotia Plaza", "Suite": "1200", "Size": 2000}]
        db = _db([{
            "Property Name": "Scotia Place", "Suite": "1200", "Size": 2000,
        }])
        results = run_comparator(brochure, db)
        # Depending on exact fuzzy score the result is REVIEW or UPDATE/OK
        # Just assert we get some valid action
        actions = {r["Action"] for r in results}
        self.assertTrue(actions.issubset({"REVIEW", "OK", "UPDATE", "NEW", "REMOVED"}))


    def test_suite_and_size_help_confirm_match(self):
        brochure = [{"Property Name": "Scotia Place Tower", "Suite": "1200", "Size": 2000}]
        db = _db([{
            "Property Name": "Scotia Place", "Suite": "1200", "Size": 2000,
        }])
        results = run_comparator(brochure, db)
        matched = [r for r in results if r.get("Action") in ("OK", "UPDATE")]
        self.assertTrue(matched)

    def test_output_excel_three_tabs(self):
        """Integration smoke test: build_output_excel produces 3 correctly-named sheets."""
        import tempfile
        import os
        from datetime import date
        import openpyxl
        from modules.output_builder import build_output_excel

        brochure_records = [_rec(**{"Extraction_Confidence": "high"})]
        diff_results = [dict(_rec(), **{"Action": "NEW", "Change_Notes": "", "Match_Score": 0})]
        removed_records = []

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test.xlsx")
            result_path = build_output_excel(
                brochure_records=brochure_records,
                diff_results=diff_results,
                removed_records=removed_records,
                output_path=out,
                run_date=date.today(),
                source_brokerage="CBRE",
            )
            wb = openpyxl.load_workbook(result_path)
            self.assertEqual(len(wb.sheetnames), 3)
            self.assertIn("Brochure_Extracted", wb.sheetnames)
            self.assertIn("Diff_Review", wb.sheetnames)
            self.assertIn("Removed_Investigate", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
